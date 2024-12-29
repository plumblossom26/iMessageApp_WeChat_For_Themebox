import os
from PIL import Image, UnidentifiedImageError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as ExcelImage


def apply_background_to_column(ws, start_row, end_row, col, color):
    """为指定列的单元格添加背景颜色"""
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=col).fill = fill


def simplify_filename_without_normal(filename):
    """移除_Normal字段的文件名"""
    return filename.replace("_Normal", "")


def process_images_to_excel(folder_path, excel_file, sheet_name="Sheet1"):
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
    else:
        wb = load_workbook(excel_file)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

    # 设置表头
    ws.cell(row=1, column=1, value="图标")
    ws.cell(row=1, column=2, value="序号")
    ws.cell(row=1, column=3, value="文件名")
    ws.cell(row=1, column=4, value="基础尺寸")
    ws.cell(row=1, column=5, value="2x尺寸")
    ws.cell(row=1, column=6, value="3x尺寸")

    processed_files = []
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        extension = os.path.splitext(filename)[1].lower()
        simplified_name = simplify_filename_without_normal(os.path.splitext(filename)[0])
        file_data = {
            "filename": filename,
            "simplified_name": simplified_name.replace("@2x", "").replace("@3x", ""),
            "original_name": simplified_name,  # 保留完整的名称（包括@2x或@3x）
            "extension": extension,
            "base_size": None,
            "2x": None,
            "3x": None,
            "path": file_path,
        }

        if extension == ".png":
            try:
                with Image.open(file_path) as img:
                    width, height = img.size
            except UnidentifiedImageError:
                print(f"无法读取文件 {filename} 的尺寸。")
                continue

            if "@2x" in filename:
                file_data["base_size"] = f"{width // 2}x{height // 2}"
                file_data["2x"] = f"{width}x{height}"
            elif "@3x" in filename:
                file_data["base_size"] = f"{width // 3}x{height // 3}"
                file_data["3x"] = f"{width}x{height}"
            else:
                file_data["base_size"] = f"{width}x{height}"

        processed_files.append(file_data)

    # 确保所有文件都被处理
    processed_files.sort(key=lambda x: x["simplified_name"])

    groups = {}
    for file_data in processed_files:
        simplified_name = file_data["simplified_name"]
        if simplified_name not in groups:
            groups[simplified_name] = {"@2x": None, "@3x": None, "original": None}
        if "@2x" in file_data["original_name"]:
            groups[simplified_name]["@2x"] = file_data
        elif "@3x" in file_data["original_name"]:
            groups[simplified_name]["@3x"] = file_data
        else:
            groups[simplified_name]["original"] = file_data

    colors = ["DCE6F1", "FDE9D9"]
    color_index = 0
    row = 2
    start_row = row

    for group_name, file_variants in groups.items():
        file_data = None
        display_name = group_name
        if file_variants["original"]:  # 原始文件优先
            file_data = file_variants["original"]
        elif file_variants["@3x"]:  # 然后选择@3x
            file_data = file_variants["@3x"]
        elif file_variants["@2x"]:  # 最后选择@2x
            file_data = file_variants["@2x"]

        if file_data:
            # 如果只有单独的@2x或@3x，保留完整名称
            if not (file_variants["@2x"] and file_variants["@3x"]) and not file_variants["original"]:
                display_name = file_data["original_name"]

            # 插入图标
            if file_data["extension"] == ".png" and os.path.exists(file_data["path"]):
                img = ExcelImage(file_data["path"])
                img.height = 40
                img.width = 40
                ws.add_image(img, f"A{row}")

            # 填写文件信息
            ws.cell(row=row, column=2, value=row - 1)  # 序号
            ws.cell(row=row, column=3, value=display_name)  # 文件名
            ws.cell(row=row, column=4, value=file_data["base_size"] if file_data["base_size"] else "/")
            ws.cell(row=row, column=5, value=file_variants["@2x"]["2x"] if file_variants["@2x"] else "/")
            ws.cell(row=row, column=6, value=file_variants["@3x"]["3x"] if file_variants["@3x"] else "/")

            # 确保图标、序号和文件名对齐
            img.anchor = f"A{row}"  # 确保图标对应的行正确
            row += 1

        apply_background_to_column(ws, start_row, row - 1, 3, colors[color_index])
        color_index = 1 - color_index
        start_row = row

    wb.save(excel_file)
    print(f"所有文件信息已成功写入到Excel文件 {excel_file} 中。")


# 使用示例
folder_path = "输入路径"  # 替换为你的文件夹路径
excel_file = "输入路径"  # 替换为你的Excel文件路径
process_images_to_excel(folder_path, excel_file)